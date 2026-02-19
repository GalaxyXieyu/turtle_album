#!/usr/bin/env python3
"""
TurtleAlbum Excel Export Script

根据用户需求导出产品数据到 Excel 文件。

Usage:
    python3 scripts/export_to_excel.py --env prod --password PASSWORD --output products.xlsx
    python3 scripts/export_to_excel.py --env prod --password PASSWORD --series CB-2026
    python3 scripts/export_to_excel.py --env prod --password PASSWORD --quality-max 7
"""

import requests
import argparse
from typing import List, Dict, Any, Optional
from datetime import datetime
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ 缺少依赖: openpyxl")
    print("请安装: pip3 install openpyxl")
    sys.exit(1)


class TurtleAlbumAPI:
    """TurtleAlbum API 客户端"""

    ENVIRONMENTS = {
        "dev": "http://localhost:8000",
        "staging": "https://staging.turtlealbum.com",
        "prod": "https://turtlealbum.com"
    }

    def __init__(self, env: str, username: str, password: str):
        if env not in self.ENVIRONMENTS:
            raise ValueError(f"Invalid environment: {env}")

        self.base_url = self.ENVIRONMENTS[env]
        self.env = env
        self.token = None
        self.login(username, password)

    def login(self, username: str, password: str):
        """登录并获取 token"""
        try:
            response = requests.post(
                f"{self.base_url}/api/auth/login",
                json={"username": username, "password": password},
                timeout=10
            )
            response.raise_for_status()
            self.token = response.json()["access_token"]
            print(f"✅ 登录成功 ({self.env})")
        except requests.exceptions.RequestException as e:
            print(f"❌ 登录失败: {e}")
            raise

    def get_headers(self) -> Dict[str, str]:
        return {
            "Authorization": f"Bearer {self.token}",
            "Content-Type": "application/json"
        }

    def get_all_products(self) -> List[Dict[str, Any]]:
        """获取所有产品"""
        try:
            response = requests.get(
                f"{self.base_url}/api/products",
                params={"page": 1, "page_size": 1000},
                headers=self.get_headers(),
                timeout=10
            )
            response.raise_for_status()
            return response.json()["products"]
        except requests.exceptions.RequestException as e:
            print(f"❌ 获取产品失败: {e}")
            raise

    def get_all_series(self) -> List[Dict[str, Any]]:
        """获取所有系列"""
        try:
            response = requests.get(
                f"{self.base_url}/api/series",
                headers=self.get_headers(),
                timeout=10
            )
            response.raise_for_status()
            return response.json()["data"]
        except requests.exceptions.RequestException as e:
            print(f"❌ 获取系列失败: {e}")
            raise


class DataQualityAnalyzer:
    """数据质量分析器"""

    @staticmethod
    def analyze_product(product: Dict[str, Any]) -> Dict[str, Any]:
        """分析产品质量"""
        score = 5
        missing_fields = []

        if product.get("description"):
            score += 1
            if len(product["description"]) >= 100:
                score += 0.5
        else:
            missing_fields.append("description")

        if product.get("images") and len(product["images"]) > 0:
            score += 1
            if len(product["images"]) >= 3:
                score += 1
        else:
            missing_fields.append("images")

        if product.get("dimensions"):
            score += 1
        else:
            missing_fields.append("dimensions")

        if product.get("series_id"):
            score += 0.5
        else:
            missing_fields.append("series_id")

        if product.get("cost_price") and product["cost_price"] > 0:
            score += 0.5
        else:
            missing_fields.append("cost_price")

        if product.get("has_sample"):
            score += 0.25

        if product.get("box_dimensions") and product.get("box_quantity"):
            score += 0.25
        else:
            if not product.get("box_dimensions"):
                missing_fields.append("box_dimensions")
            if not product.get("box_quantity"):
                missing_fields.append("box_quantity")

        if score >= 9:
            level = "excellent"
        elif score >= 7:
            level = "good"
        elif score >= 5:
            level = "fair"
        else:
            level = "poor"

        return {
            "score": round(score, 1),
            "level": level,
            "missing_fields": missing_fields
        }


class ExcelExporter:
    """Excel 导出器"""

    def __init__(self, filename: str):
        self.filename = filename
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)  # 删除默认 sheet

    def create_products_sheet(self, products: List[Dict[str, Any]], series_map: Dict[str, str]):
        """创建产品数据表"""
        ws = self.workbook.create_sheet("Products")

        # 定义列
        headers = [
            "货号", "产品名称", "产品描述", "形状", "材质",
            "出厂价格", "成本价", "重量(kg)", "长度(cm)", "宽度(cm)", "高度(cm)",
            "容量最小(ml)", "容量最大(ml)", "分隔数量",
            "产品类型", "管型", "盒型", "工艺类型", "功能设计",
            "库存状态", "有样品", "是否精选", "热度评分",
            "纸箱尺寸", "装箱数量",
            "系列编号", "系列名称",
            "性别", "后代单价", "父本编号", "母本编号",
            "图片数量", "主图URL",
            "质量评分", "质量等级", "缺失字段"
        ]

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 写入数据
        for row_idx, product in enumerate(products, 2):
            analysis = DataQualityAnalyzer.analyze_product(product)
            dims = product.get("dimensions") or {}
            capacity = dims.get("capacity") or {}
            images = product.get("images") or []
            main_image = next((img for img in images if img.get("type") == "main"), None)

            series_name = series_map.get(product.get("series_id"), "")

            data = [
                product.get("code", ""),
                product.get("name", ""),
                product.get("description", ""),
                product.get("shape", ""),
                product.get("material", ""),
                product.get("factory_price", 0),
                product.get("cost_price", 0),
                dims.get("weight", ""),
                dims.get("length", ""),
                dims.get("width", ""),
                dims.get("height", ""),
                capacity.get("min", ""),
                capacity.get("max", ""),
                dims.get("compartments", ""),
                product.get("product_type", ""),
                product.get("tube_type", ""),
                product.get("box_type", ""),
                product.get("process_type", ""),
                product.get("functional_designs", ""),
                "有货" if product.get("in_stock") else "缺货",
                "是" if product.get("has_sample") else "否",
                "是" if product.get("is_featured") else "否",
                product.get("popularity_score", 0),
                product.get("box_dimensions", ""),
                product.get("box_quantity", ""),
                product.get("series_id", ""),
                series_name,
                product.get("sex", ""),
                product.get("offspring_unit_price", ""),
                product.get("sire_code", ""),
                product.get("dam_code", ""),
                len(images),
                main_image.get("url") if main_image else "",
                analysis["score"],
                analysis["level"],
                ", ".join(analysis["missing_fields"])
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = Alignment(vertical="center")

                # 质量评分着色
                if col == len(data) - 2:  # 质量评分列
                    if analysis["score"] >= 9:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif analysis["score"] >= 7:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif analysis["score"] < 5:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # 冻结首行
        ws.freeze_panes = "A2"

    def create_images_sheet(self, products: List[Dict[str, Any]]):
        """创建图片数据表"""
        ws = self.workbook.create_sheet("Images")

        headers = ["产品编号", "产品名称", "图片类型", "图片URL", "图片描述", "排序"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_idx = 2
        for product in products:
            images = product.get("images") or []
            for image in images:
                data = [
                    product.get("code", ""),
                    product.get("name", ""),
                    image.get("type", ""),
                    image.get("url", ""),
                    image.get("alt", ""),
                    image.get("sort_order", 0)
                ]

                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.alignment = Alignment(vertical="center")

                row_idx += 1

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        ws.freeze_panes = "A2"

    def create_quality_report_sheet(self, products: List[Dict[str, Any]]):
        """创建质量报告表"""
        ws = self.workbook.create_sheet("Quality Report")

        # 分析所有产品
        scores = []
        distribution = {"excellent": 0, "good": 0, "fair": 0, "poor": 0}
        missing_fields_count = {}

        for product in products:
            analysis = DataQualityAnalyzer.analyze_product(product)
            scores.append(analysis["score"])
            distribution[analysis["level"]] += 1

            for field in analysis["missing_fields"]:
                missing_fields_count[field] = missing_fields_count.get(field, 0) + 1

        total = len(products)
        avg_score = sum(scores) / len(scores) if scores else 0

        # 写入报告
        row = 1

        # 标题
        ws.cell(row=row, column=1, value="数据质量报告").font = Font(size=16, bold=True)
        row += 2

        # 总体统计
        ws.cell(row=row, column=1, value="总体统计").font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="总产品数:")
        ws.cell(row=row, column=2, value=total)
        row += 1
        ws.cell(row=row, column=1, value="平均评分:")
        ws.cell(row=row, column=2, value=f"{avg_score:.2f}/10")
        row += 2

        # 质量分布
        ws.cell(row=row, column=1, value="质量分布").font = Font(bold=True)
        row += 1
        for level, label in [("excellent", "优秀 (9-10分)"), ("good", "良好 (7-9分)"),
                              ("fair", "一般 (5-7分)"), ("poor", "较差 (0-5分)")]:
            count = distribution[level]
            percentage = (count / total * 100) if total > 0 else 0
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
            row += 1

        row += 1

        # 最常缺失的字段
        ws.cell(row=row, column=1, value="最常缺失的字段").font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="字段名").font = Font(bold=True)
        ws.cell(row=row, column=2, value="缺失数量").font = Font(bold=True)
        ws.cell(row=row, column=3, value="缺失比例").font = Font(bold=True)
        row += 1

        sorted_missing = sorted(missing_fields_count.items(), key=lambda x: x[1], reverse=True)
        for field, count in sorted_missing[:15]:
            percentage = (count / total * 100) if total > 0 else 0
            ws.cell(row=row, column=1, value=field)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
            row += 1

        # 调整列宽
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

    def save(self):
        """保存文件"""
        self.workbook.save(self.filename)
        print(f"✅ 文件已保存: {self.filename}")


def filter_products(products: List[Dict[str, Any]], args) -> List[Dict[str, Any]]:
    """根据参数筛选产品"""
    filtered = products

    # 按系列筛选
    if args.series:
        filtered = [p for p in filtered if p.get("series_id") == args.series or
                    any(s.get("code") == args.series or s.get("name") == args.series
                        for s in [p.get("series")] if s)]

    # 按产品类型筛选
    if args.product_type:
        filtered = [p for p in filtered if p.get("product_type") == args.product_type]

    # 按库存状态筛选
    if args.in_stock is not None:
        filtered = [p for p in filtered if p.get("in_stock") == args.in_stock]

    # 按质量评分筛选
    if args.quality_min is not None or args.quality_max is not None:
        def check_quality(p):
            analysis = DataQualityAnalyzer.analyze_product(p)
            score = analysis["score"]
            if args.quality_min is not None and score < args.quality_min:
                return False
            if args.quality_max is not None and score > args.quality_max:
                return False
            return True

        filtered = [p for p in filtered if check_quality(p)]

    # 按缺失字段筛选
    if args.missing_field:
        def has_missing_field(p):
            analysis = DataQualityAnalyzer.analyze_product(p)
            return args.missing_field in analysis["missing_fields"]

        filtered = [p for p in filtered if has_missing_field(p)]

    return filtered


def main():
    parser = argparse.ArgumentParser(description="TurtleAlbum Excel 导出工具")
    parser.add_argument("--env", choices=["dev", "staging", "prod"], default="dev",
                        help="环境")
    parser.add_argument("--username", default="admin", help="用户名")
    parser.add_argument("--password", required=True, help="密码")
    parser.add_argument("--output", default=f"turtle_album_products_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        help="输出文件名")

    # 筛选选项
    parser.add_argument("--series", help="按系列筛选")
    parser.add_argument("--product-type", choices=["tube", "box", "turtle"], help="按产品类型筛选")
    parser.add_argument("--in-stock", type=bool, help="按库存状态筛选")
    parser.add_argument("--quality-min", type=float, help="最低质量评分")
    parser.add_argument("--quality-max", type=float, help="最高质量评分")
    parser.add_argument("--missing-field", help="筛选缺失指定字段的产品")

    args = parser.parse_args()

    # 初始化 API
    api = TurtleAlbumAPI(args.env, args.username, args.password)

    # 获取数据
    print("⏳ 正在获取产品数据...")
    products = api.get_all_products()
    print(f"✅ 获取到 {len(products)} 个产品")

    print("⏳ 正在获取系列数据...")
    series_list = api.get_all_series()
    series_map = {s["id"]: s["name"] for s in series_list}
    print(f"✅ 获取到 {len(series_list)} 个系列")

    # 筛选产品
    filtered_products = filter_products(products, args)
    print(f"✅ 筛选后: {len(filtered_products)} 个产品")

    if not filtered_products:
        print("⚠️ 没有符合条件的产品")
        return

    # 导出到 Excel
    print(f"⏳ 正在生成 Excel 文件...")
    exporter = ExcelExporter(args.output)
    exporter.create_products_sheet(filtered_products, series_map)
    exporter.create_images_sheet(filtered_products)
    exporter.create_quality_report_sheet(filtered_products)
    exporter.save()

    print(f"\n📊 导出完成!")
    print(f"文件: {args.output}")
    print(f"产品数: {len(filtered_products)}")


if __name__ == "__main__":
    main()
